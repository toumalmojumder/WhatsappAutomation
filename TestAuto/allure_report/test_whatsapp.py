import time
import openpyxl
import pyperclip
from selenium import webdriver
import allure
import pytest
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

@allure.severity(allure.severity_level.NORMAL)
class TestWhatApp:
    @allure.severity(allure.severity_level.NORMAL)
    def test_search(self):
        self.driver = webdriver.Chrome("C:\\chromedriver_win32\\chromedriver.exe")
        self.driver.get("https://web.whatsapp.com/")
        excel_path = "J:\\Automation\\WhatsappAutomation\\WhatsappAutomation\\TestAuto\\contacts.xlsx"
        time.sleep(5)
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active

        search_xpath = '//div[@contenteditable="true"][@data-tab="3"]'
        search_box = WebDriverWait(self.driver, 20).until(
            EC.presence_of_element_located((By.XPATH, search_xpath))
        )

        pyperclip.copy(sheet.cell(1, 1).value)
        search_box.send_keys(Keys.CONTROL + "v")

        contact_xpath = f'//span[@title="{sheet.cell(1, 1).value}"]'
        contact_title = WebDriverWait(self.driver, 20).until(
            EC.presence_of_element_located((By.XPATH, contact_xpath))
        )

        contact_title.click()
        time.sleep(1)
        self.driver.close()
        assert True

    @allure.severity(allure.severity_level.NORMAL)
    def test_send_message(self):
        self.driver = webdriver.Chrome("C:\\chromedriver_win32\\chromedriver.exe")
        excel_path = "J:\\Automation\\WhatsappAutomation\\WhatsappAutomation\\TestAuto\\contacts.xlsx"

        self.driver.get("https://web.whatsapp.com/")

        time.sleep(5)

        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active

        rows = sheet.max_row

        for r in range(1, rows + 1):
            search_xpath = '//div[@contenteditable="true"][@data-tab="3"]'
            search_box = WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((By.XPATH, search_xpath))
            )
            search_box.clear()

            time.sleep(1)

            pyperclip.copy(sheet.cell(r, 1).value)
            search_box.send_keys(Keys.CONTROL + "v")

            time.sleep(1)

            contact_xpath = f'//span[@title="{sheet.cell(r, 1).value}"]'
            contact_title = WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((By.XPATH, contact_xpath))
            )

            contact_title.click()
            time.sleep(1)

            input_xpath = '//div[@contenteditable="true"][@data-tab="9"]'
            input_box = WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((By.XPATH, input_xpath))
            )

            pyperclip.copy(sheet.cell(r, 2).value)
            input_box.send_keys(Keys.CONTROL + "v")
            input_box.send_keys(Keys.ENTER)
            time.sleep(1)
        self.driver.close()
        assert True

    @allure.severity(allure.severity_level.NORMAL)
    def test_excel(self):
        self.driver = webdriver.Chrome("C:\\chromedriver_win32\\chromedriver.exe")
        excel_path = "J:\\Automation\\WhatsappAutomation\\WhatsappAutomation\\TestAuto\\contacts.xlsx"

        self.driver.get("https://web.whatsapp.com/")

        time.sleep(5)

        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active
        rows = sheet.max_row

        for r in range(1, rows + 1):
            search_xpath = '//div[@contenteditable="true"][@data-tab="3"]'
            search_box = WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((By.XPATH, search_xpath))
            )
            search_box.clear()

            time.sleep(1)

            pyperclip.copy(sheet.cell(r, 1).value)
            search_box.send_keys(Keys.CONTROL + "v")

            time.sleep(1)

            contact_xpath = f'//span[@title="{sheet.cell(r, 1).value}"]'
            contact_title = WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((By.XPATH, contact_xpath))
            )

            contact_title.click()
            time.sleep(1)

            input_xpath = '//div[@contenteditable="true"][@data-tab="9"]'
            input_box = WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((By.XPATH, input_xpath))
            )

            pyperclip.copy(sheet.cell(r, 2).value)
            input_box.send_keys(Keys.CONTROL + "v")
            input_box.send_keys(Keys.ENTER)
            time.sleep(1)
            sheet.cell(r, 3).value = "Sent"
            workbook.save(excel_path)
            workbook.close()
        self.driver.close()
        assert True

    @allure.severity(allure.severity_level.NORMAL)
    def test_logout(self):
        self.driver = webdriver.Chrome("C:\\chromedriver_win32\\chromedriver.exe")
        excel_path = "J:\\Automation\\WhatsappAutomation\\WhatsappAutomation\\TestAuto\\contacts.xlsx"

        self.driver.get("https://web.whatsapp.com/")

        time.sleep(5)

        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active

        rows = sheet.max_row

        for r in range(1, rows + 1):
            search_xpath = '//div[@contenteditable="true"][@data-tab="3"]'
            search_box = WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((By.XPATH, search_xpath))
            )
            search_box.clear()

            time.sleep(1)

            pyperclip.copy(sheet.cell(r, 1).value)
            search_box.send_keys(Keys.CONTROL + "v")

            time.sleep(1)

            contact_xpath = f'//span[@title="{sheet.cell(r, 1).value}"]'
            contact_title = WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((By.XPATH, contact_xpath))
            )

            contact_title.click()
            time.sleep(1)

            input_xpath = '//div[@contenteditable="true"][@data-tab="9"]'
            input_box = WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((By.XPATH, input_xpath))
            )

            pyperclip.copy(sheet.cell(r, 2).value)
            input_box.send_keys(Keys.CONTROL + "v")
            input_box.send_keys(Keys.ENTER)
            time.sleep(1)
            sheet.cell(r, 3).value = "Sent"
            time.sleep(1)
            workbook.save(excel_path)
            workbook.close()
        time.sleep(3)
        # logout
        menu_1_xpath = '//div[@tabindex="0"] [@data-tab="6"] [@title="Menu"] '
        menu_1_btn = WebDriverWait(self.driver, 20).until(
            EC.presence_of_element_located((By.XPATH, menu_1_xpath))
        )
        menu_1_btn.click()
        time.sleep(2)
        close_xpath = '//li[@tabindex="-1"] //div[@aria-label="Close chat"]'
        close_btn = WebDriverWait(self.driver, 20).until(
            EC.presence_of_element_located((By.XPATH, close_xpath))
        )
        close_btn.click()
        time.sleep(2)
        menu_xpath = '//span[@data-testid="menu"][@data-icon="menu"]'
        menu_btn = WebDriverWait(self.driver, 20).until(
            EC.presence_of_element_located((By.XPATH, menu_xpath))
        )
        menu_btn.click()
        time.sleep(2)

        logout_xpath = '//li[@tabindex="-1"] //div[@aria-label = "Log out"]'
        logout_btn = WebDriverWait(self.driver, 20).until(
            EC.presence_of_element_located((By.XPATH, logout_xpath))
        )
        logout_btn.click()
        self.driver.close()
        assert True



