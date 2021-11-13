import time
import openpyxl
import pyperclip

from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


driver = webdriver.Chrome("C:\\chromedriver_win32\\chromedriver.exe")
excel_path = "J:\\Automation\\WhatsappAutomation\\WhatsappAutomation\\TestAuto\\contacts.xlsx"

driver.get("https://web.whatsapp.com/")

time.sleep(5)

workbook = openpyxl.load_workbook(excel_path)
sheet = workbook.active

rows = sheet.max_row

for r in range(1, rows+1):
    search_xpath = '//div[@contenteditable="true"][@data-tab="3"]'
    search_box = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, search_xpath))
    )
    search_box.clear()

    time.sleep(1)

    pyperclip.copy(sheet.cell(r, 1).value)
    search_box.send_keys(Keys.CONTROL+ "v")

    time.sleep(1)

    contact_xpath = f'//span[@title="{sheet.cell(r, 1).value}"]'
    contact_title = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, contact_xpath))
    )
    # contact_title = driver.find_element_by_xpath(contact_xpath)
    contact_title.click()
    time.sleep(1)

    input_xpath = '//div[@contenteditable="true"][@data-tab="9"]'
    input_box = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, input_xpath))
    )
    # input_box = driver.find_element_by_xpath(input_xpath)

    pyperclip.copy(sheet.cell(r, 2).value)
    input_box.send_keys(Keys.CONTROL+ "v")
    input_box.send_keys(Keys.ENTER)
    time.sleep(1)
    sheet.cell(r, 3).value = "Sent"
    workbook.save(excel_path)
    workbook.close()

print("Successfully write result on excel")
