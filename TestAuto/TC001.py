import time
import openpyxl
import pyperclip

from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

driver = webdriver.Chrome("C:\\chromedriver_win32\\chromedriver.exe")
excel_path = "J:\\Automation\\WhatsappAutomation\\WhatsappAutomation\\TestAuto\\contacts.xlsx"

driver.get("https://web.whatsapp.com/")

time.sleep(5)

workbook = openpyxl.load_workbook(excel_path)
sheet = workbook.active

rows = sheet.max_row

search_xpath = '//div[@contenteditable="true"][@data-tab="3"]'
search_box = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, search_xpath))
)

pyperclip.copy(sheet.cell(1, 1).value)
search_box.send_keys(Keys.CONTROL, Keys.v)
print("Display searched contact")

